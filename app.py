import openpyxl, sys # type: ignore

path = "test.xlsx"

wb_obj = openpyxl.load_workbook(path)

sheet_obj = wb_obj.active

row = sheet_obj.max_row
column = sheet_obj.max_column

arr = []
for c in range(2, row + 1):
    c_obj = sheet_obj.cell(row=c, column=1)
    arr.append(c_obj.value)
    
if len(sys.argv) == 1:
    print(arr)
else:
    c = sheet_obj.cell(row=int(sys.argv[1]), column=2)
    c.value = "OK"
    wb_obj.save(path)
